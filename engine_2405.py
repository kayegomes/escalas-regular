import os
import re
import unicodedata

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from app_support import append_execution_history, load_app_config, save_with_fallback
from engine_grades import process_all_grades


def _normalize_text(value):
    if pd.isna(value):
        return ""
    text = str(value).strip().upper()
    if not text or text in {"NAN", "NONE", "NAT", "-"}:
        return ""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _normalize_channel(value):
    text = _normalize_text(value)
    if not text:
        return ""

    aliases = {
        "SPORTV": "SPORTV",
        "SPORTV 1": "SPORTV",
        "SPORTV1": "SPORTV",
        "SPORTV 2": "SPORTV2",
        "SPORTV2": "SPORTV2",
        "SPORTV 3": "SPORTV3",
        "SPORTV3": "SPORTV3",
        "SPORTV 4": "SPORTV4",
        "SPORTV4": "SPORTV4",
        "PREMIERE": "PREMIERE",
        "PREMIERE CLUBES": "PREMIERE",
        "TV GLOBO": "TV GLOBO",
        "TV GLOBO REDE": "TV GLOBO",
        "GLOBO": "TV GLOBO",
        "COMBATE": "COMBATE",
    }

    for key, canonical in aliases.items():
        if key in text:
            return canonical

    return text


def _parse_date(value):
    dt = pd.to_datetime(value, errors="coerce", dayfirst=True)
    if pd.isna(dt):
        return pd.NaT
    return dt.normalize()


def _parse_series_date(value):
    if pd.isna(value):
        return pd.NaT
    dt = pd.to_datetime(value, errors="coerce", dayfirst=True)
    if pd.isna(dt):
        return pd.NaT
    return dt.normalize()


def _channel_matches(grade_platform, wo_channel):
    grade_platform = _normalize_channel(grade_platform)
    wo_channel = _normalize_channel(wo_channel)

    if not grade_platform or not wo_channel:
        return True

    if grade_platform == wo_channel:
        return True

    if grade_platform == "SPORTV" and wo_channel.startswith("SPORTV"):
        return True

    return False


def _event_score(grade_event, wo_row):
    grade_norm = _normalize_text(grade_event)
    if not grade_norm:
        return 0.0

    candidates = [
        wo_row.get("W/O Description", ""),
        wo_row.get("Event Group", ""),
        wo_row.get("Job Description", ""),
        wo_row.get("Spec Note", ""),
        wo_row.get("Episode Description", ""),
    ]

    best = 0.0
    for candidate in candidates:
        wo_norm = _normalize_text(candidate)
        if not wo_norm:
            continue

        if grade_norm == wo_norm:
            return 1.0

        if grade_norm in wo_norm or wo_norm in grade_norm:
            ratio = min(len(grade_norm), len(wo_norm)) / max(len(grade_norm), len(wo_norm))
            best = max(best, 0.92 * ratio + 0.08)
            continue

        grade_tokens = set(grade_norm.split())
        wo_tokens = set(wo_norm.split())
        if not grade_tokens or not wo_tokens:
            continue

        overlap = len(grade_tokens & wo_tokens) / len(grade_tokens | wo_tokens)
        best = max(best, overlap)

    return best


def _load_2405_report(file_path):
    df = pd.read_excel(file_path)
    df.columns = [str(c).strip() for c in df.columns]

    if "Contact" in df.columns:
        contact = df["Contact"].astype(str).str.strip()
        df = df[contact.isin(["", "nan", "NaN", "None"])]

    if "W/O Description" not in df.columns:
        raise ValueError("Coluna 'W/O Description' não encontrada no relatório 2405.")

    if "Air Begin" in df.columns:
        df["_air_begin_date"] = df["Air Begin"].apply(_parse_date)
    elif "W/O Begin" in df.columns:
        df["_air_begin_date"] = df["W/O Begin"].apply(_parse_date)
    else:
        df["_air_begin_date"] = pd.NaT

    df["_channel_norm"] = df["Channel"].apply(_normalize_channel) if "Channel" in df.columns else ""
    df["_event_norm"] = df["W/O Description"].apply(_normalize_text)
    df["_group_norm"] = df["Event Group"].apply(_normalize_text) if "Event Group" in df.columns else ""
    return df


def run_etapa1_2405(path_2405, path_sp1, path_sp2, path_pr1, path_pr2, path_co1, path_co2):
    """
    Etapa 1 do fluxo:
    cruza os eventos válidos das grades com o relatório 2405 de WOs independentes.
    """
    if not path_2405 or not os.path.exists(path_2405):
        raise FileNotFoundError("Selecione o relatório 2405 antes de executar a checagem.")

    df_2405 = _load_2405_report(path_2405)
    df_grades = process_all_grades(path_sp1, path_sp2, path_pr1, path_pr2, path_co1, path_co2)

    if df_grades.empty:
        raise ValueError("Nenhum evento válido foi encontrado nas grades fornecidas.")

    df_grades = df_grades.copy()
    df_grades["Data_obj"] = df_grades.get("Data").apply(_parse_series_date)
    df_grades = df_grades[df_grades["Data_obj"].notna()].copy()
    df_grades["_channel_norm"] = df_grades["Plataforma"].apply(_normalize_channel)
    df_grades["_event_norm"] = df_grades["Evento"].apply(_normalize_text)
    df_grades = df_grades[df_grades["_event_norm"] != ""].copy()

    if "V/I" in df_grades.columns:
        df_grades = df_grades[~df_grades["V/I"].astype(str).str.upper().isin(["R", "REPRISE"])].copy()

    df_grades["_date_key"] = df_grades["Data_obj"].dt.strftime("%Y-%m-%d")
    df_grades = df_grades.drop_duplicates(subset=["_date_key", "_channel_norm", "_event_norm"]).copy()

    resultados = []
    for _, grade_row in df_grades.iterrows():
        grade_date = grade_row["Data_obj"].normalize()
        grade_channel = grade_row["_channel_norm"]
        grade_event = grade_row["Evento"]

        candidates = df_2405.copy()
        if pd.notna(grade_date) and "_air_begin_date" in candidates.columns:
            same_date = candidates["_air_begin_date"] == grade_date
            if same_date.any():
                candidates = candidates[same_date].copy()

        if not candidates.empty and grade_channel:
            channel_mask = candidates["_channel_norm"].apply(lambda ch: _channel_matches(grade_channel, ch))
            filtered = candidates[channel_mask].copy()
            if not filtered.empty:
                candidates = filtered

        best_row = None
        best_score = 0.0
        for _, wo_row in candidates.iterrows():
            score = _event_score(grade_event, wo_row)
            if score > best_score:
                best_score = score
                best_row = wo_row

        status = "AUSENTE"
        obs = "Não encontrado no 2405"
        wo_number = ""
        wo_phase = ""
        wo_channel = ""
        wo_desc = ""
        match_field = ""

        if best_row is not None and best_score >= 0.55:
            wo_number = str(best_row.get("Work Order", "")).strip()
            wo_phase = str(best_row.get("Phase", "")).strip()
            wo_channel = str(best_row.get("Channel", "")).strip()
            wo_desc = str(best_row.get("W/O Description", "")).strip()
            status = "OK"
            obs = "Encontrado no 2405"
            match_field = "W/O Description"
            if best_score < 0.85:
                status = "PARCIAL"
                obs = "Encontrado com correspondência parcial"

        resultados.append(
            {
                "Status Checagem": status,
                "Observação": obs,
                "Score": round(best_score, 3),
                "WO#": wo_number,
                "Phase": wo_phase,
                "Channel 2405": wo_channel,
                "W/O Description 2405": wo_desc,
                "Campo Casado": match_field,
                "Plataforma": grade_row.get("Plataforma", ""),
                "Data": grade_row.get("Data", ""),
                "Início": grade_row.get("Início", ""),
                "Pré": grade_row.get("Pré", ""),
                "Fim": grade_row.get("Fim", ""),
                "Evento": grade_event,
                "V/I": grade_row.get("V/I", ""),
            }
        )

    df_out = pd.DataFrame(resultados)

    output_name = load_app_config().get("outputs", {}).get("etapa_1", "Check_2405_Gerado.xlsx")
    out_path = os.path.join(os.path.dirname(path_2405), output_name)

    def write_output(target_path):
        with pd.ExcelWriter(target_path, engine="openpyxl") as writer:
            df_out.to_excel(writer, index=False, sheet_name="Checagem")
            resumo = (
                df_out.groupby("Status Checagem", dropna=False)
                .size()
                .reset_index(name="Quantidade")
                .sort_values("Quantidade", ascending=False)
            )
            resumo.to_excel(writer, index=False, sheet_name="Resumo")

    out_path, used_fallback = save_with_fallback(write_output, out_path)

    wb = load_workbook(out_path)
    ws = wb["Checagem"]

    red_fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFFFCC", end_color="FFFFFFCC", fill_type="solid")
    green_fill = PatternFill(start_color="FFC6E0B4", end_color="FFC6E0B4", fill_type="solid")
    red_font = Font(color="FF9C0006", bold=True)
    yellow_font = Font(color="FF9C6500", bold=True)
    green_font = Font(color="FF006100", bold=True)

    header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    status_col = header_map.get("Status Checagem")
    if status_col:
        for row_idx in range(2, ws.max_row + 1):
            status = str(ws.cell(row=row_idx, column=status_col).value or "")
            if status == "OK":
                ws.cell(row=row_idx, column=status_col).fill = green_fill
                ws.cell(row=row_idx, column=status_col).font = green_font
            elif status == "PARCIAL":
                ws.cell(row=row_idx, column=status_col).fill = yellow_fill
                ws.cell(row=row_idx, column=status_col).font = yellow_font
            else:
                ws.cell(row=row_idx, column=status_col).fill = red_fill
                ws.cell(row=row_idx, column=status_col).font = red_font

    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col_idx).value
        if not header:
            continue
        max_len = len(str(header))
        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 3, 60)

    wb.save(out_path)
    append_execution_history(
        os.path.dirname(path_2405),
        "Etapa 1 - Checagem 2405 vs grades",
        "SUCESSO",
        f"{len(df_out)} eventos processados; saída: {os.path.basename(out_path)}",
        "Arquivo alternativo utilizado por bloqueio do Excel." if used_fallback else "",
    )
    return out_path
