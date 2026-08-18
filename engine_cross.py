import os
import unicodedata
from datetime import timedelta

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app_support import append_execution_history, load_app_config, save_with_fallback
from engine_2468 import process_2468_base
from engine_grades import process_all_grades


def _norm_text(value):
    text = str(value if value is not None else "").strip().upper()
    text = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def _find_column(row, candidates, default=""):
    for key in candidates:
        for col in row.index:
            if _norm_text(col) == _norm_text(key):
                value = row.get(col)
                if pd.notna(value):
                    return value
    return default


def _row_text(row):
    parts = [
        _find_column(row, ["Tipo de Atividade", "Tipo Atividade"]),
        _find_column(row, ["Row Display"]),
        _find_column(row, ["Sub-Atividade", "Sub-Atividade (shift)"]),
        _find_column(row, ["Descrição", "Evento/Programa", "Atividade/Descrição", "Evento"]),
        _find_column(row, ["Produto (WO/Quick Hold)", "Produto (WO/Shift)"]),
        _find_column(row, ["Quick Hold Job Info"]),
        _find_column(row, ["Quick Hold Job Details"]),
        _find_column(row, ["Event Group"]),
    ]
    return " | ".join(_norm_text(part) for part in parts)


def _is_missing_pre(value):
    if pd.isna(value):
        return True
    return _norm_text(value) in {"", "-", "X", "NAN", "NAT", "NONE"}


def _is_travel_row(row):
    return "VIAGEM" in _row_text(row)


def _is_folga_row(row):
    combined = _row_text(row)
    if "VIAGEM" in combined:
        return False
    return any(token in combined for token in ["FOLGA", "DAY OFF", "VACATION", "FERIAS", "COMP DAY"])


def _is_quickhold_in_scale(row):
    tipo = _norm_text(_find_column(row, ["Tipo de Atividade", "Tipo Atividade"]))
    if "QUICK HOLD" not in tipo:
        return False

    combined = _row_text(row)
    if "VIAGEM" in combined:
        return False
    if "AUSENCIA MEDICA" in combined or ("AUSENCIA" in combined and "MEDICA" in combined):
        return False

    return (
        "PODCAST" in combined
        or "CABINE DO JOGO" in combined
        or ("PARTICIPACAO" in combined and "CABINE" in combined and "JOGO" in combined)
    )


def _is_valid_time_str(val):
    if val is None or pd.isna(val):
        return False
    text = str(val).strip()
    if not text or text.upper() in {"", "-", "NAN", "NAT", "NONE"}:
        return False
    if any(bad in text.upper() for bad in ["#", "DEFINIR", "DEF", "UNDEF"]):
        return False
    parts = text.split(":")
    if len(parts) >= 2 and parts[0].isdigit() and parts[1].isdigit():
        return True
    return False


def _platform_match_mask(df_grades, plat_2468, relaxed=False):
    plat = _norm_text(plat_2468)
    series = df_grades["Plataforma"].astype(str).str.upper()

    # GE.com é uma frente digital vinculada à programação do Sportv. GE TV
    # permanece fora deste fluxo, pois possui envio separado.
    if "GE.COM" in plat or "GECOM" in plat:
        return series.str.contains("SPORTV1|SPORTV 1|^SPORTV$", na=False)
    if "SPORTV" in plat:
        if relaxed:
            return series.str.contains("SPORTV", na=False)
        if "2" in plat:
            return series.str.contains("SPORTV2|SPORTV 2", na=False)
        elif "3" in plat:
            return series.str.contains("SPORTV3|SPORTV 3", na=False)
        elif "1" in plat or plat in {"SPORTV", "SPORTV 1", "SPORTV1"}:
            return series.str.contains("SPORTV1|SPORTV 1|^SPORTV$", na=False)
        return series.str.contains("SPORTV", na=False)
    if "PREMIERE" in plat:
        return series == "PREMIERE"
    if "COMBATE" in plat:
        return series == "COMBATE"
    if "GLOBO" in plat:
        return series == "TV GLOBO"
    if plat:
        return series.str.contains(plat[:4], na=False)
    return pd.Series(True, index=df_grades.index)


def _grade_date_key(value):
    if pd.isna(value):
        return ""
    try:
        dt = pd.to_datetime(value, errors="coerce", dayfirst=True)
        if pd.notna(dt):
            return dt.strftime("%Y-%m-%d")
    except Exception:
        pass
    text = str(value).strip()
    return text[:10] if len(text) >= 10 else text


def _event_search_text(row_2468):
    parts = [
        _find_column(row_2468, ["Atividade/Descrição", "Evento/Programa", "Descrição", "Evento"]),
        _find_column(row_2468, ["Row Display"]),
        _find_column(row_2468, ["Sub-Atividade", "Sub-Atividade (shift)"]),
        _find_column(row_2468, ["Produto (WO/Quick Hold)", "Produto (WO/Shift)"]),
        _find_column(row_2468, ["Event Group"]),
    ]
    return " ".join(str(p).strip() for p in parts if p is not None and str(p).strip() not in {"", "nan", "NaT"})


GENERIC_WORDS = {
    "LIGA", "DAS", "NACOES", "NAÇOES", "DE", "VOLEI", "VÔLEI", "MASCULINA", "FEMININA",
    "PRIMEIRA", "FASE", "VT", "EVENTO", "AO", "VIVO", "RODADA", "ETAPA", "JOGO", "MATCH",
    "COPA", "DO", "MUNDO", "GRAND", "PRIX", "SÉRIE", "SERIE", "A", "B", "C", "D"
}


def _parse_time_seconds(time_val):
    if pd.isna(time_val):
        return None
    text = str(time_val).strip()
    if not text or text.upper() in {"NONE", "NAN", "NAT", "-"}:
        return None
    try:
        dt = pd.to_datetime(text, errors="coerce")
        if pd.notna(dt):
            return dt.hour * 3600 + dt.minute * 60 + dt.second
    except Exception:
        pass
    parts = text.split(":")
    if len(parts) >= 2 and parts[0].isdigit() and parts[1].isdigit():
        return int(parts[0]) * 3600 + int(parts[1]) * 60
    return None


def _score_grade_match(g_row, ev_norm, time_2468_sec=None):
    g_ev = _norm_text(g_row.get("Evento"))
    if not g_ev:
        return -100

    g_words = set(w for w in g_ev.split() if len(w) > 1)
    ev_words = set(w for w in ev_norm.split() if len(w) > 1)

    common_words = g_words.intersection(ev_words)
    if not common_words:
        return -100

    generic_words_norm = {_norm_text(word) for word in GENERIC_WORDS}
    meaningful_common_words = common_words - generic_words_norm
    # Não aceitar uma linha genérica como "VT DE EVENTO" apenas porque o
    # relatório também contém a palavra genérica EVENTO. Isso criava matches
    # falsos em datas sem o evento específico na grade.
    if not meaningful_common_words and g_ev != ev_norm:
        return -100

    score = 0
    for w in common_words:
        if w not in GENERIC_WORDS:
            score += 15
        else:
            score += 1

    if g_ev == ev_norm:
        score += 50
    elif g_ev in ev_norm or ev_norm in g_ev:
        score += 20

    g_inicio_sec = _parse_time_seconds(g_row.get("Início"))
    g_pre_sec = _parse_time_seconds(g_row.get("Pré"))

    if time_2468_sec is not None:
        best_diff = None
        for t_sec in [g_inicio_sec, g_pre_sec]:
            if t_sec is not None:
                diff = abs(time_2468_sec - t_sec)
                diff = min(diff, 86400 - diff)
                if best_diff is None or diff < best_diff:
                    best_diff = diff
        if best_diff is not None:
            diff_min = best_diff / 60.0
            if diff_min <= 15:
                score += 30
            elif diff_min <= 60:
                score += 20
            elif diff_min <= 180:
                score += 10
            else:
                score -= min(30, int(diff_min / 60) * 5)

    return score


def _match_event_in_pool(search_pool, ev_norm, time_2468_sec=None, min_score=1):
    if search_pool.empty or not ev_norm:
        return None

    best_score = -100
    best_row = None
    for _, g_row in search_pool.iterrows():
        s = _score_grade_match(g_row, ev_norm, time_2468_sec)
        if s > best_score:
            best_score = s
            best_row = g_row

    if best_score >= min_score:
        return best_row
    return None


def _grade_text(row):
    return _norm_text(_find_column(row, ["Evento", "EVENTO"]))


def _mark_missing_grade(out_row, alertas, severity):
    """Sinaliza que os horários não foram confirmados na grade.

    Os horários originais do relatório permanecem na saída para permitir a
    conferência, mas nunca devem ser apresentados como horários confirmados
    quando não existe uma linha correspondente na grade.
    """

    alertas.append("Horário não encontrado na Grade")
    if severity == "OK":
        severity = "YELLOW"
    out_row["Pré"] = "-"
    return severity


def _append_pre_review_alerts(out_row, alertas, severity, is_prog, has_valid_grade_time):
    """Aplica a regra de Pré para programas e eventos esportivos.

    Programas podem não ter Pré por natureza. Eventos esportivos, porém,
    precisam ser conferidos quando possuem Início, mas não possuem Pré válido,
    mesmo que a grade marque a coluna Pré como ``X``.
    """

    if is_prog or not has_valid_grade_time:
        return severity

    val_pre = out_row.get("Pré")
    val_inicio = out_row.get("Início")
    tem_pre = not _is_missing_pre(val_pre)
    tem_inicio = pd.notnull(val_inicio) and str(val_inicio).strip() not in {"", "nan", "NaT", "None"}

    if tem_inicio and not tem_pre:
        alertas.append("Conferir Pré")
        if severity == "OK":
            severity = "YELLOW"
    elif tem_inicio and tem_pre and _norm_text(val_pre) == _norm_text(val_inicio):
        alertas.append("Pré igual ao Início")
        if severity == "OK":
            severity = "YELLOW"

    return severity


def _format_date_value(value):
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if text in {"", "-", "nan", "NaT", "None"}:
        return ""
    try:
        dt = pd.to_datetime(value, errors="coerce", dayfirst=not text.startswith("20"))
        if pd.isna(dt):
            return text.split(" ")[0]
        return dt.strftime("%d/%m/%Y")
    except Exception:
        return text.split(" ")[0]


def _format_time_value(value):
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if text in {"", "-", "nan", "NaT", "None"}:
        return ""
    try:
        dt = pd.to_datetime(value, errors="coerce", dayfirst=not text.startswith("20"))
        if pd.notna(dt):
            return dt.strftime("%H:%M")
    except Exception:
        pass
    if " " in text:
        text = text.split(" ")[-1]
    parts = text.split(":")
    return f"{parts[0]}:{parts[1]}" if len(parts) >= 2 else text


def _normalize_output_columns(df_out):
    rename_map = {}
    for col in df_out.columns:
        norm = _norm_text(col)
        if norm == "NOME":
            rename_map[col] = "Nome"
        elif norm == "TIPO ATIVIDADE":
            rename_map[col] = "Tipo Atividade"
        elif norm == "FUNCAO":
            rename_map[col] = "Função"
        elif norm == "REPTER":
            rename_map[col] = "Repórter"
        elif norm == "PRE":
            rename_map[col] = "Pré"
        elif norm == "INICIO":
            rename_map[col] = "Início"
        elif norm == "FIM":
            rename_map[col] = "Fim"
        elif norm == "STATUS REVISAO":
            rename_map[col] = "Status Revisão"
        elif norm == "LOCAL NARRACAO":
            rename_map[col] = "Local Narração"
        elif norm == "LOCAL DE GRAVACAO":
            rename_map[col] = "Local de Gravação"
        elif norm == "ATIVIDADE/DESCRICAO":
            rename_map[col] = "Atividade/Descrição"
        elif norm == "DESCRICAO":
            rename_map[col] = "Descrição"
        elif norm == "TIPO DE PRODUCAO":
            rename_map[col] = "Tipo de Produção"
        elif norm == "HR PRODUCAO":
            rename_map[col] = "Hr Produção"
    if rename_map:
        df_out = df_out.rename(columns=rename_map)
    return df_out


def _is_programa(row_2468, best_match=None):
    tipo_prod = _norm_text(_find_column(row_2468, ["Tipo de Produção"]))
    if any(k in tipo_prod for k in ["ESTUDIO", "INTERNET"]):
        return True

    # "VT no Controle" representa conteúdo gravado. A coluna V/I da grade
    # diferencia conteúdo transmissivo/inédito (V ou I) de programas/blocos
    # sem indicação de transmissão. Para o caso I, mantemos a ausência de Pré
    # sinalizada enquanto a regra de negócio ainda está em observação.
    if "VT" in tipo_prod:
        if best_match is None:
            return True
        grade_vi = _norm_text(_find_column(best_match, ["V/I", "VI", "V I"]))
        return grade_vi not in {"V", "I", "AO VIVO", "LIVE", "AO VIVO/INEDITO"}

    tipo_ativ = _norm_text(_find_column(row_2468, ["Tipo de Atividade", "Tipo Atividade"]))
    if "QUICK HOLD" in tipo_ativ:
        return True

    combined = " | ".join([
        _norm_text(_find_column(row_2468, ["Evento/Programa", "Descrição", "Atividade/Descrição", "Evento"])),
        _norm_text(_find_column(row_2468, ["Produto (WO/Quick Hold)", "Produto (WO/Shift)"])),
        _norm_text(_find_column(row_2468, ["Event Group"]))
    ])

    program_keywords = [
        "BDRJ", "BOM DIA", "GLOBO ESPORTE", "GE ", "GE -", "ESPORTE ESPETACULAR", "EE ", "EE -",
        "SELECAO", "SPORTV NEWS", "TROCA DE PASSE", "REDACAO", "TROCANDO PASSE", "COMBATE NEWS",
        "BASTIDORES", "PROGRAMA", "ROUND", "PODCAST", "GRID DA FORMULA"
    ]
    if any(kw in combined for kw in program_keywords):
        return True

    if best_match is not None:
        grade_ev = _norm_text(_find_column(best_match, ["Evento", "EVENTO"]))
        if any(kw in grade_ev for kw in program_keywords):
            return True

    return False


def _find_best_grade_match(row_2468, df_grades):
    plat_2468 = str(_find_column(row_2468, ["Plataforma", "Canal (Master Room)", "Canal"])).strip().upper()
    if plat_2468 == "NAN":
        plat_2468 = ""

    evento_2468 = _event_search_text(row_2468)
    data_2468 = row_2468.get("Data_raw", row_2468.get("Data"))

    dt_2468 = None
    dt_2468_str = ""
    time_2468_sec = None
    report_hour = None
    if pd.notna(data_2468):
        try:
            data_text = str(data_2468).strip()
            is_brazilian_date = "/" in data_text and not data_text.startswith("20")
            dt_2468 = pd.to_datetime(data_2468, errors="coerce", dayfirst=is_brazilian_date)
            if pd.notna(dt_2468):
                dt_2468_str = dt_2468.strftime("%Y-%m-%d")
                report_hour = dt_2468.hour
                time_2468_sec = dt_2468.hour * 3600 + dt_2468.minute * 60 + dt_2468.second
        except Exception:
            pass

    # A coluna Data da Base 2468 frequentemente contém apenas a data (DD/MM/YYYY).
    # Nesse caso, a hora real vem de Início/Air Start Time e não deve ser tratada como 00:00.
    report_start = _find_column(row_2468, ["Início", "Inicio", "Air Start Time", "Hora de Início"])
    report_start_sec = _parse_time_seconds(report_start)
    if report_start_sec is not None:
        time_2468_sec = report_start_sec
        report_hour = report_start_sec // 3600

    ev_norm = _norm_text(evento_2468)
    if not ev_norm:
        return None

    match = df_grades
    if plat_2468:
        match = match[_platform_match_mask(match, plat_2468)]
        if match.empty:
            match = df_grades[_platform_match_mask(df_grades, plat_2468, relaxed=True)]

    if match.empty:
        return None

    date_keys = match["Data"].apply(_grade_date_key) if "Data" in match.columns else pd.Series("", index=match.index)

    def _pool_for_dates(keys):
        if not keys:
            return match
        mask = pd.Series(False, index=match.index)
        for key in keys:
            if key:
                mask = mask | (date_keys == key)
        return match[mask] if mask.any() else pd.DataFrame()

    search_sequence = []
    if dt_2468_str:
        adjacent = [dt_2468_str]
        if dt_2468 is not None:
            if report_hour is not None and report_hour < 6:
                adjacent.append((dt_2468 - timedelta(days=1)).strftime("%Y-%m-%d"))
            elif report_hour is not None and report_hour >= 21:
                adjacent.append((dt_2468 + timedelta(days=1)).strftime("%Y-%m-%d"))
        search_sequence.append(_pool_for_dates(adjacent))
    if not dt_2468_str:
        search_sequence.append(match)
    else:
        # Nunca usar uma ocorrência distante apenas porque o título coincide.
        # A grade pode ter o mesmo evento em vários dias; fora da própria data
        # (ou da tolerância de virada de madrugada) o horário não está confirmado.
        no_date_pool = match[date_keys == ""]
        if not no_date_pool.empty:
            search_sequence.append(no_date_pool)

    seen = set()
    for pool in search_sequence:
        if pool.empty:
            continue
        pool_id = id(pool)
        if pool_id in seen:
            continue
        seen.add(pool_id)

        min_s = 1 if pool is match else 5
        result = _match_event_in_pool(pool, ev_norm, time_2468_sec=time_2468_sec, min_score=min_s)
        if result is not None:
            return result

    return None


def run_etapa1(path_2468, path_sp1, path_sp2, path_pr1, path_pr2, path_co1, path_co2):
    print("Iniciando Motor 2468...")
    df_2468 = process_2468_base(path_2468)

    print("Iniciando Motor de Grades...")
    df_grades = process_all_grades(path_sp1, path_sp2, path_pr1, path_pr2, path_co1, path_co2)

    print("Iniciando Cruzamento (Match)...")

    if not df_2468.empty:
        df_2468 = df_2468[~df_2468.apply(_is_travel_row, axis=1)]

    if not df_2468.empty and any(_norm_text(c) == "FUNCAO" for c in df_2468.columns):
        total_antes = len(df_2468)
        funcao_col = next(c for c in df_2468.columns if _norm_text(c) == "FUNCAO")
        funcao_lower = df_2468[funcao_col].astype(str).str.strip().str.lower()
        mask_talento = funcao_lower.str.contains("narrador", na=False) | funcao_lower.str.contains("coment", na=False)
        mask_folga = df_2468.apply(_is_folga_row, axis=1)
        mask_quickhold = df_2468.apply(_is_quickhold_in_scale, axis=1)
        df_2468 = df_2468[mask_talento | mask_folga | mask_quickhold]
        removidos = total_antes - len(df_2468)
        print(
            f"Filtro de elenco aplicado: {removidos} linhas removidas "
            f"(Coordenadores, Produtores, Repórteres, etc.). Restaram {len(df_2468)} linhas."
        )

    output_rows = []

    for _, row_2468 in df_2468.iterrows():
        out_row = row_2468.to_dict()
        alertas = []
        severity = "OK"

        evento_2468 = str(_find_column(row_2468, ["Evento/Programa", "Descrição", "Atividade/Descrição", "Evento"])).strip()
        plat_2468 = str(_find_column(row_2468, ["Plataforma", "Canal (Master Room)", "Canal"])).strip().upper()
        if plat_2468 == "NAN":
            plat_2468 = ""

        if _is_folga_row(row_2468):
            out_row["Status Revisão"] = "OK"
            out_row["_severity"] = "OK"
            output_rows.append(out_row)
            continue

        if _is_quickhold_in_scale(row_2468):
            local_narr = str(_find_column(row_2468, ["Local de Locução", "Local Narração", "Local"])).strip().upper()
            if local_narr in {"", "NAN", "NAT", "NONE"}:
                alertas.append("Local Ausente")
                severity = "YELLOW"
            out_row["Pré"] = "-"
            out_row["Status Revisão"] = "OK" if not alertas else " | ".join(alertas)
            out_row["_severity"] = severity
            output_rows.append(out_row)
            continue

        if df_grades.empty:
            alertas.append("Sem Grades Fornecidas")
            severity = "RED"
            out_row["Status Revisão"] = " | ".join(alertas)
            out_row["_severity"] = severity
            output_rows.append(out_row)
            continue

        best_match = _find_best_grade_match(row_2468, df_grades)
        is_prog = _is_programa(row_2468, best_match)
        has_valid_grade_time = False

        if best_match is not None:
            if any(token in _norm_text(evento_2468) for token in ["SURF", "TENIS"]):
                alertas.append("Fallback (Multimodalidade)")
                if severity == "OK":
                    severity = "YELLOW"
            else:
                inicio_grade = _find_column(best_match, ["Início", "InÇðcio", "Inicio"])
                fim_grade = _find_column(best_match, ["Fim", "FIM"])
                pre_grade = _find_column(best_match, ["Pré", "Pr", "Pré ", "PrÇ¸"])

                inicio_fmt = _format_time_value(inicio_grade)
                fim_fmt = _format_time_value(fim_grade)
                pre_fmt = _format_time_value(pre_grade)

                if not _is_valid_time_str(inicio_fmt):
                    severity = _mark_missing_grade(out_row, alertas, severity)
                else:
                    has_valid_grade_time = True
                    out_row["Início"] = inicio_fmt
                    # Grade Fim is calculated as the start of the next V/R event in the schedule,
                    # which is the actual end of the broadcast block. Use it when available;
                    # the 2468 Fim acts as fallback only when grade has no Fim.
                    if fim_fmt and _is_valid_time_str(fim_fmt):
                        out_row["Fim"] = fim_fmt

                    if is_prog or _norm_text(pre_grade) == "X" or _norm_text(pre_fmt) in {"X", "-", "", "NAN", "NAT", "NONE"}:
                        out_row["Pré"] = "-"
                    else:
                        out_row["Pré"] = pre_fmt

            if "CONFIRMAR" in _grade_text(best_match):
                alertas.append("A Confirmar")
                if severity == "OK":
                    severity = "YELLOW"
        else:
            # Não encontrado na grade pelo título. Os horários do relatório
            # ficam visíveis como referência, mas o status sempre sinaliza que
            # não houve confirmação, inclusive quando a atividade é programa.
            severity = _mark_missing_grade(out_row, alertas, severity)

        severity = _append_pre_review_alerts(
            out_row,
            alertas,
            severity,
            is_prog,
            has_valid_grade_time,
        )

        tipo_atividade = _norm_text(_find_column(row_2468, ["Tipo de Atividade", "Tipo Atividade"]))
        if tipo_atividade in {"BOOKING", "QUICK HOLD"}:
            local_narr = str(_find_column(row_2468, ["Local de Locução", "Local Narração", "Local"])).strip().upper()
            if local_narr in {"", "NAN", "NAT", "NONE"}:
                alertas.append("Local Ausente")
                if severity == "OK":
                    severity = "YELLOW"

        out_row["Status Revisão"] = "OK" if not alertas else " | ".join(alertas)
        out_row["_severity"] = severity
        output_rows.append(out_row)

    df_out = pd.DataFrame(output_rows)

    severity_map = {}
    if "_severity" in df_out.columns:
        severity_map = df_out["_severity"].to_dict()
        df_out = df_out.drop(columns=["_severity"])

    if "Data_raw" in df_out.columns:
        df_out = df_out.drop(columns=["Data_raw"])

    if "Data" in df_out.columns:
        df_out["Data"] = df_out["Data"].apply(_format_date_value)
    for col in ["Pré", "Início", "Fim"]:
        if col in df_out.columns:
            df_out[col] = df_out[col].apply(_format_time_value)

    df_out = _normalize_output_columns(df_out)

    preferred_order = [
        "Nome",
        "Tipo Atividade",
        "Descrição",
        "Row Display",
        "Sub-Atividade (shift)",
        "Resource Phase",
        "Cliente",
        "Canal (Master Room)",
        "Data",
        "Dia",
        "Pré",
        "Início",
        "Fim",
        "WO#",
        "Hr Produção",
        "WO Status",
        "Tipo de Produção",
        "Produto (WO/Shift)",
        "Quick Hold Job Info",
        "Quick Hold Job Details",
        "Event Group",
        "Atividade/Descrição",
        "Equipe",
        "Função",
        "Local Narração",
        "notas",
        "Local de Gravação",
        "Cidade",
        "UF",
        "Status",
        "Parent #",
        "Plataforma",
        "Narrador",
        "Comentarista",
        "Repórter",
        "Elenco",
        "Coordenador",
        "Produtor",
    ]
    ordered_cols = [c for c in preferred_order if c in df_out.columns]
    ordered_cols.extend([c for c in df_out.columns if c not in ordered_cols and c != "Status Revisão"])
    if "Status Revisão" in df_out.columns:
        ordered_cols.append("Status Revisão")
    df_out = df_out[ordered_cols]

    output_name = load_app_config().get("outputs", {}).get("etapa_2", "Check_Pre_Envio_Gerado.xlsx")
    out_path = os.path.join(os.path.dirname(path_2468), output_name)
    out_path, used_fallback = save_with_fallback(
        lambda target_path: df_out.to_excel(target_path, index=False),
        out_path,
    )

    wb = load_workbook(out_path)
    ws = wb.active

    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
    green_fill = PatternFill(start_color="FF00B050", end_color="FF00B050", fill_type="solid")

    red_font = Font(color="FFFF0000", bold=True)
    yellow_font = Font(color="FF9C6500", bold=True)
    green_font = Font(color="FF006100", bold=True)
    white_font = Font(color="FFFFFFFF", bold=True)

    status_col_idx = None
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col_idx).value
        if header and "STATUS REVISAO" in _norm_text(header):
            status_col_idx = col_idx
            break

    if status_col_idx:
        for row_idx in range(2, ws.max_row + 1):
            status_cell = ws.cell(row=row_idx, column=status_col_idx)
            sev = severity_map.get(row_idx - 2, "OK")

            if sev == "RED":
                status_cell.fill = red_fill
                status_cell.font = white_font
                light_red = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
                for col in range(1, ws.max_column + 1):
                    if col != status_col_idx:
                        ws.cell(row=row_idx, column=col).fill = light_red
            elif sev == "YELLOW":
                status_cell.fill = yellow_fill
                status_cell.font = yellow_font
                light_yellow = PatternFill(start_color="FFFFFFCC", end_color="FFFFFFCC", fill_type="solid")
                for col in range(1, ws.max_column + 1):
                    if col != status_col_idx:
                        ws.cell(row=row_idx, column=col).fill = light_yellow
            else:
                status_cell.fill = green_fill
                status_cell.font = white_font

        max_len = len("Status Revisão")
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=status_col_idx).value
            if val and len(str(val)) > max_len:
                max_len = len(str(val))
        ws.column_dimensions[ws.cell(row=1, column=status_col_idx).column_letter].width = min(max_len + 4, 60)

    # Legenda em uma aba própria para não ocupar a área operacional da tabela.
    if "Legenda" in wb.sheetnames:
        del wb["Legenda"]
    legend_ws = wb.create_sheet("Legenda")
    legend_ws.sheet_view.showGridLines = False
    legend_ws.merge_cells("A1:B1")
    legend_ws["A1"] = "Legenda — Status Revisão"
    legend_ws["A1"].font = Font(bold=True, size=14, color="FFFFFFFF")
    legend_ws["A1"].fill = PatternFill(start_color="FF1F4E78", end_color="FF1F4E78", fill_type="solid")
    legend_ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    legend_ws.row_dimensions[1].height = 28

    legend_ws["A3"] = "Status Revisão"
    legend_ws["B3"] = "Significado"
    header_fill = PatternFill(start_color="FFD9EAF7", end_color="FFD9EAF7", fill_type="solid")
    border = Border(
        left=Side(style="thin", color="FFD9D9D9"),
        right=Side(style="thin", color="FFD9D9D9"),
        top=Side(style="thin", color="FFD9D9D9"),
        bottom=Side(style="thin", color="FFD9D9D9"),
    )
    for cell in legend_ws[3]:
        cell.font = Font(bold=True, color="FF1F1F1F")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    status_legend = [
        ("OK", "Correspondência encontrada e sem alerta pendente. Programas sem Pré podem permanecer OK quando localizados na grade."),
        ("Conferir Pré", "Evento ou conteúdo transmissivo encontrado, mas sem horário de Pré válido. A conferência manual é necessária."),
        ("Pré igual ao Início", "O Pré e o horário de Início ficaram iguais; conferir se o Pré foi lançado corretamente na grade."),
        ("Horário não encontrado na Grade", "Nenhuma correspondência confiável foi encontrada. Os horários exibidos permanecem apenas como referência do relatório."),
        ("Fallback (Multimodalidade)", "Correspondência encontrada em modalidade com maior risco de variação de nomenclatura, como surfe ou tênis; revisar manualmente."),
        ("A Confirmar", "A grade encontrou o evento, mas o próprio registro está marcado como a confirmar."),
        ("Local Ausente", "A atividade não possui local de locução preenchido no relatório."),
        ("Sem Grades Fornecidas", "Não foi fornecida uma grade válida para executar o cruzamento."),
    ]
    legend_fills = {
        "OK": PatternFill(start_color="FF00B050", end_color="FF00B050", fill_type="solid"),
        "Horário não encontrado na Grade": PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid"),
        "Sem Grades Fornecidas": PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid"),
        "Conferir Pré": PatternFill(start_color="FFFFFFCC", end_color="FFFFFFCC", fill_type="solid"),
        "Pré igual ao Início": PatternFill(start_color="FFFFE699", end_color="FFFFE699", fill_type="solid"),
        "Fallback (Multimodalidade)": PatternFill(start_color="FFFCE4D6", end_color="FFFCE4D6", fill_type="solid"),
        "A Confirmar": PatternFill(start_color="FFFFFFCC", end_color="FFFFFFCC", fill_type="solid"),
        "Local Ausente": PatternFill(start_color="FFFFFFCC", end_color="FFFFFFCC", fill_type="solid"),
    }
    for row_idx, (status, meaning) in enumerate(status_legend, start=4):
        status_cell = legend_ws.cell(row=row_idx, column=1, value=status)
        meaning_cell = legend_ws.cell(row=row_idx, column=2, value=meaning)
        status_cell.fill = legend_fills[status]
        status_cell.font = Font(bold=True, color="FF1F1F1F")
        for cell in (status_cell, meaning_cell):
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        legend_ws.row_dimensions[row_idx].height = 42

    note_row = 4 + len(status_legend) + 2
    legend_ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=2)
    legend_ws.cell(row=note_row, column=1, value='Quando houver mais de um alerta, os status aparecem combinados e separados por " | ".')
    legend_ws.cell(row=note_row, column=1).font = Font(italic=True, color="FF666666")
    legend_ws.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True)
    legend_ws.column_dimensions["A"].width = 34
    legend_ws.column_dimensions["B"].width = 105
    legend_ws.freeze_panes = "A4"
    legend_ws.auto_filter.ref = f"A3:B{3 + len(status_legend)}"
    legend_ws.sheet_properties.pageSetUpPr.fitToPage = True
    legend_ws.page_setup.fitToWidth = 1
    legend_ws.page_setup.fitToHeight = 0

    wb.save(out_path)
    append_execution_history(
        os.path.dirname(path_2468),
        "Etapa 2 - Cruzamento 2468 vs grades",
        "SUCESSO",
        f"{len(df_out)} linhas processadas; saída: {os.path.basename(out_path)}",
        "Arquivo alternativo utilizado por bloqueio do Excel." if used_fallback else "",
    )
    return out_path
